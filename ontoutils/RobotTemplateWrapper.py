import csv
import logging
import re
from typing import Optional

import openpyxl
from openpyxl import Workbook

from ontoutils.core.ColumnMapping import ColumnMapping, get_relationship_mapping, DEFAULT_HEADER_MAPPINGS, \
    DEFAULT_HEADERS_TO_IGNORE, RobotType
from ontoutils.RobotWrapper import RobotWrapper
from ontoutils.core import OntologyEntity, OntologyRelation
from ontoutils.utils import quoteIfNeeded, quoted


def patch_entity_from_excel_col(entity: OntologyEntity, col_val: str, header_val: ColumnMapping):
    if header_val.robotType == RobotType.ROBOT_TYPE_ID:
        entity.id = col_val

    if header_val.robotType == RobotType.ROBOT_TYPE_LABEL:
        name = col_val
        entity.name = name
        entity.synonyms = []
        if '(' in name and ')' in name:
            synonym = re.search(r'\((.*?)\)', name).group(1)
            if len(synonym) > 0:
                name = name[:name.index("(")].strip()
                entity.name = name
                entity.synonyms = [synonym]

    if header_val.excelColName == "Synonyms":
        more_synonyms = col_val.split(";")
        entity.synonyms.extend(more_synonyms)

    if header_val.excelColName == "Definition":
        entity.definition = col_val

    if header_val.excelColName == "Parent":
        parents = col_val
        parent = parents.split("/")[0]
        if '(' in parent:
            parent = parent[:parent.index("(")].strip()
        if '[' in parent:
            parent = parent[:parent.index("[")].strip()
        entity.parent = parent

    if header_val.excelColName == "Examples":
        examples = col_val
        if len(examples) > 0:
            entity.examples = examples

    if header_val.excelColName == "Comment":
        comment = col_val
        if len(comment) > 0:
            entity.comment = comment

    if header_val.excelColName == "Curation status":
        entity.curationStatus = col_val

class RobotTemplateWrapper(RobotWrapper):
    _logger = logging.getLogger(__name__)

    all_entity_names: dict[str, OntologyEntity]
    '''
    index of names to ontology entities
    '''

    all_entity_ids: dict[str, OntologyEntity]
    '''
    index of ids to ontology entities
    '''

    all_rel_names: dict[str, OntologyRelation]
    '''
    index of names to relations
    '''

    all_rel_ids: dict[str, OntologyRelation]
    '''
    index of ids to relations
    '''

    parents_to_children: dict[str, list[str]]

    header_mapping: dict[str, ColumnMapping]

    ignored_headers: list[str]

    def __init__(self, robotcmd):
        super().__init__(robotcmd, True)
        self.all_entity_names = {}
        self.all_entity_ids = {}
        self.all_rel_names = {}
        self.all_rel_ids = {}
        self.parents_to_children = {}
        self.header_mapping = DEFAULT_HEADER_MAPPINGS
        self.ignored_headers = DEFAULT_HEADERS_TO_IGNORE

    def __dfs__(self, order, node):
        if node not in order.keys() and node.lower() in self.all_entity_names.keys():
            order[node] = ''
            if node in self.parents_to_children.keys():
                for child in self.parents_to_children[node]:
                    self.__dfs__(order, child)
        return order

    def add_classes_from_excel(self, excel_file_name: str, csv_file_name: Optional[str] = None) -> None:
        """
        Adds classes and their relations from an excel file. Optionally writes a robot template csv file.

        if the csv_file_name parameter is supplied the robot template csv file will be written at that path.

        :param excel_file_name: Path to the exec file
        :param csv_file_name: Path to the output csv file
        :return:
        """

        write_csv = csv_file_name is not None

        try:
            wb = openpyxl.load_workbook(excel_file_name)
        except Exception as e:
            self._logger.error(f"Failed to open excel sheet '{excel_file_name}': {e}")
            raise Exception("Error! Not able to parse file: " + excel_file_name)

        sheet = wb.active
        data = sheet.rows

        headers = self._extract_headers_for_class_def(data, excel_file_name)

        headers_mapped = [self.header_mapping[h] for h in headers if h is not None and h not in self.ignored_headers]
        header_indices = [i for i, h in enumerate(headers) if h is not None and h not in self.ignored_headers]

        # Process the rows, create a CSV template at the same time
        if write_csv:
            csvfile = open(csv_file_name, 'w', newline='')
            csv_writer = csv.writer(csvfile, delimiter=',', quotechar='\"', quoting=csv.QUOTE_MINIMAL)

            csv_writer.writerow([headers[i] for i in header_indices])
            csv_writer.writerow([c.get_robot_code_string() for c in headers_mapped])

        for raw_row in data:
            row: list[Optional[str]] = [raw_row[i].value for i in
                                        header_indices]  # just those headers that are mapped
            row_with_header: list[tuple[Optional[str], ColumnMapping]] = list(zip(row, headers_mapped))
            new_row: list[str] = [mapping.parse_value(i) for (i, mapping) in row_with_header]

            entity = OntologyEntity()
            # Now also process and store the values for merging if needed
            for value, mapping in row_with_header:
                if value is None:
                    continue

                patch_entity_from_excel_col(entity, value, mapping)

            self.all_entity_ids[entity.id] = entity
            self.all_entity_names[entity.name.lower()] = entity
            self.all_entity_names[entity.name.lower()] = entity
            for synonym in entity.synonyms:
                self.all_entity_names[synonym.lower()] = entity

            if write_csv:
                if entity.curationStatus not in ['Obsolete']:
                    csv_writer.writerow(new_row)
                else:
                    self._logger.info(f"Not writing row for entity '{entity.name}' to template due to obsolete status")

        if write_csv:
            csvfile.close()

        self._logger.debug('FINISHED PARSING ALL ROWS IN SPREADSHEET')
        wb.close()

    def _extract_headers_for_class_def(self, data, excel_file_name: str):
        header: list[str] = [i.value for i in next(data)]
        self._logger.debug(f"Headers for '{excel_file_name}': {header}")

        # Check all header strings are in the header mapping or else fail with an error
        headers_not_mapped = [h for h in header if h not in self.header_mapping.keys()]
        self._logger.debug(f"Headers initially not mapped for '{excel_file_name}': {headers_not_mapped}")

        for h in headers_not_mapped:
            if h is not None and h.strip().startswith('REL'):
                values = quoted.findall(h)
                if len(values) == 1:
                    self.header_mapping[h] = get_relationship_mapping(h, rel_id=quoteIfNeeded(values[0]))
                else:
                    self._logger.warning(f"Relation column did not match expected format. Value: '{h}'")

        headers_not_mapped = [h for h in header if
                              h not in self.header_mapping.keys() and h is not None and h not in self.ignored_headers]

        if len(headers_not_mapped) > 0:
            self._logger.warning(f"Headers not mapped for '{excel_file_name}': {headers_not_mapped}. Ignoring ...")
            for h in headers_not_mapped:
                self.ignored_headers.append(h)
        return header

    def add_rel_info_from_excel(self, excel_file_name: str) -> None:
        """
        Adds relation

        :param excel_file_name:
        :return:
        """

        try:
            wb = openpyxl.load_workbook(excel_file_name)
        except Exception as e:
            self._logger.error(f"Failed to open excel sheet '{excel_file_name}': {e}")
            raise Exception("Error! Not able to parse file: " + excel_file_name)

        sheet = wb.active
        data = sheet.rows

        header = [i.value for i in next(data)[0:7]]
        self._logger.debug(header)

        for row in data:
            rowdata: list[str] = [i.value for i in row[0:7]]
            id = rowdata[0]
            name = rowdata[1]

            if name is None:
                continue

            entity = OntologyRelation(id, name)
            entity.equivalent = rowdata[2]
            entity.parent = rowdata[3]
            entity.definition = rowdata[4]
            entity.domain = rowdata[5]
            entity.range = rowdata[6]

            self.all_rel_names[name.lower()] = entity
            self.all_rel_ids[id] = entity

        wb.close()

    def create_csv_relation_template_file(self, csv_file_name: str):
        # Create ROBOT template for NEW properties (parent is not None)

        with open(csv_file_name, 'w', newline='') as rel_create_csv:
            rel_writer = csv.writer(rel_create_csv, delimiter=',', quotechar='\"', quoting=csv.QUOTE_MINIMAL)
            rel_writer.writerow(["Id", "Name", "Type", "Parent", "Def", "Domain", "Range"])
            rel_writer.writerow(["ID", "LABEL", "TYPE", "SP %", "A IAO:0000115", "DOMAIN", "RANGE"])

            for rel in self.all_rel_names.values():
                parent = rel.parent[rel.parent.find("[") + 1:rel.parent.find("]")] if rel.parent is not None else None
                rel_writer.writerow([rel.id, rel.name, "object property", parent, rel.domain, rel.range])

        self._logger.debug(f"Finished writing relation creation template at '{csv_file_name}'")

    def mergeRelInfoFromLucidChart(self, entities, relations):
        # Merge lucidchart information with definitions information to populate relations
        # for entity in entities.values():
        #    entity_name = entity.name.replace('\n',' ') # just the name matters
        #    # do we have this entity name in our all_entities?
        #    if entity_name.lower().strip() not in self.all_entity_names.keys():
        #        print("ERROR: Name ", entity_name, "NOT FOUND.")
        #        print(self.all_entity_names.keys())

        for rel in relations:
            rel_name = rel.relType
            if '(' in rel_name:
                rel_name = rel_name[0:rel_name.rindex('(')].strip()
            if rel_name.lower() not in self.all_rel_names.keys():
                print("ERROR: Rel ", rel_name, "NOT FOUND.")
                print(self.all_rel_names.keys())
                continue
            onto_rel = self.all_rel_names[rel_name.lower()]
            onto_entity1 = self.all_entity_names[rel.entity1.name.replace('\n', ' ').lower().strip()]
            onto_entity2 = self.all_entity_names[rel.entity2.name.replace('\n', ' ').lower().strip()]
            if onto_entity1.relations is None:
                onto_entity1.relations = {}
            if onto_rel.name not in onto_entity1.relations.keys():
                onto_entity1.relations[onto_rel.name] = []
            onto_entity1.relations[onto_rel.name].append(onto_entity2)

    def write_spreadsheet(self, excel_file_name, id_col_name: str) -> None:
        book = Workbook()
        sheet = book.active

        rel_header_ids = ["REL '" + s.name + "' [" + s.id + "]" for s in self.all_rel_names.values()]
        header = (id_col_name, 'Name', 'Parent', 'Definition', 'Synonyms', 'Examples', *rel_header_ids)

        sheet.append(header)

        # PARENT classes AND TARGETS OF RELATIONS -- prepare list of required imports for information and cross-checking
        import_classes = []
        top_level = []
        for entity in self.all_entity_names.values():
            parent = entity.parent
            if parent.lower() not in self.all_entity_names.keys() and parent.lower() not in import_classes:
                import_classes.append(parent.lower())
                top_level.append(entity.name)
            if parent not in self.parents_to_children:
                self.parents_to_children[parent] = []
            self.parents_to_children[parent].append(entity.name)

        self._logger.debug(f"Classes identified as imported for '{excel_file_name}': {import_classes}")

        order = {}
        for t in top_level:
            order1 = self.__dfs__({}, t)
            order.update(order1)
        for t in self.parents_to_children.keys():
            order1 = self.__dfs__({}, t)
            order.update(order1)
        order.update({x.name: '' for x in self.all_entity_ids.values()})
        order = order.keys()

        for entity_name in order:
            if entity_name is None:
                continue

            entity = self.all_entity_names[entity_name.lower()]
            parent_name = entity.parent.lower() if entity.parent.lower() in import_classes else self.all_entity_names[
                entity.parent.lower()].name
            rel_vals = [";".join([z.name for z in x]) if len(x) > 0 else '' for x in
                        [entity.relations[y.name] if entity.relations is not None and y.name in entity.relations else []
                         for y in self.all_rel_names.values()]]
            # print([y.name for y in self.all_rel_names.values()])
            # print(entity_name,rel_vals)
            row = (entity.id,
                   entity.name,
                   parent_name,
                   entity.definition,
                   ";".join(entity.synonyms),
                   entity.examples,
                   *rel_vals
                   )
            sheet.append(row)

        book.save(excel_file_name)

    # Executes ROBOT from a template file as created
    def createOntologyFromTemplateFile(self, csvFileName, dependency, iri_prefix, id_prefixes, ontology_iri,
                                       owlFileName):
        robot_cmd = [self.robotcmd, 'template', '--template', csvFileName]
        for p in id_prefixes:
            robot_cmd.append('--prefix')
            robot_cmd.append(p)
        robot_cmd.extend(['--ontology-iri', ontology_iri,
                          '--output', owlFileName
                          ])

        # A bit of hacking to deal appropriately with external dependency files:
        if dependency is not None:
            # Allow multiple dependencies. These will become OWL imports.
            dependencyFileNames = dependency.split(',')
            print(dependencyFileNames)
            dependencyFileName = "imports.owl"
            with open(dependencyFileName, 'w') as outFile:
                outFile.write("""<?xml version=\"1.0\"?>
    <rdf:RDF xmlns="http://www.semanticweb.org/ontologies/temporary#"
        xml:base="http://www.semanticweb.org/ontologies/temporary"
        xmlns:dc="http://purl.org/dc/elements/1.1/"
        xmlns:obo="http://purl.obolibrary.org/obo/"
        xmlns:owl="http://www.w3.org/2002/07/owl#"
        xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#"
        xmlns:xml="http://www.w3.org/XML/1998/namespace"
        xmlns:xsd="http://www.w3.org/2001/XMLSchema#"
        xmlns:foaf="http://xmlns.com/foaf/0.1/"
        xmlns:rdfs="http://www.w3.org/2000/01/rdf-schema#">
        <owl:Ontology rdf:about=\"""" + ontology_iri + "\">\n")
                for d in dependencyFileNames:
                    outFile.write("<owl:imports rdf:resource=\"" + iri_prefix + d + "\"/> \n")
                outFile.write(" </owl:Ontology> \n</rdf:RDF> ")

        robot_cmd.extend(['--input', dependencyFileName, "--merge-before", "--collapse-import-closure", "false"])

        robot_cmd = " ".join(robot_cmd)

        self._execute_command(command_str=robot_cmd)
