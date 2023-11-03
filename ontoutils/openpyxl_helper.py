import logging

from openpyxl import load_workbook


def open_workbook(filename: str, read_only=False, data_only=False, keep_links=True, rich_text=False):
    return ReadOnlyWorkbook(filename, read_only, data_only, keep_links, rich_text)


class ReadOnlyWorkbook:
    _logger = logging.getLogger(__name__)

    def __init__(self, file_name, read_only, data_only, keep_links, rich_text):
        self._rich_text = rich_text
        self._keep_links = keep_links
        self._data_only = data_only
        self._read_only = read_only
        self._file_name = file_name

    def __enter__(self):
        try:
            self._wb = load_workbook(self._file_name, read_only=self._read_only, data_only=self._data_only,
                                     keep_links=self._keep_links, rich_text=self._rich_text)
        except Exception as e:
            self._logger.error(f"Failed to open excel sheet '{self._file_name}': {e}")
            raise Exception("Error! Not able to parse file: " + self._file_name)

        return self._wb

    def __exit__(self, *args):
        self._wb.close()
