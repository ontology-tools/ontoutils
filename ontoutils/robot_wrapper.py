import os
import subprocess
import openpyxl
from openpyxl import Workbook
import shutil
import re
import csv
from ontoutils.core import *


quoted = re.compile("(?<=')[^']+(?=')")


def quoteIfNeeded(value):
    # first check if already existing quotes
    if not (value[0]=="'" and value[-1]=="'"):
        if " " in value:
            return( "\'"+value+"\'")
    return(value)

def getIdMapping(columnName):
    return ColumnMapping(columnName,ColumnMapping.ROBOT_TYPE_ID)
def getLabelMapping(columnName):
    return ColumnMapping(columnName,ColumnMapping.ROBOT_TYPE_LABEL)
def getParentMapping(columnName):
    return ColumnMapping(columnName,ColumnMapping.ROBOT_TYPE_PARENT)
def getRelationshipMapping(columnName,relId = None):
    if relId is None:
        relId = columnName
    return ColumnMapping(columnName,ColumnMapping.ROBOT_TYPE_RELATION,relId)
def getAnnotationMapping(columnName,annoId):
    return ColumnMapping(columnName,ColumnMapping.ROBOT_TYPE_ANNOTATION,annoId)
def getDisjointMapping(columnName):
    return ColumnMapping(columnName,ColumnMapping.ROBOT_TYPE_DISJOINT)
def getEquivalenceMapping(columnName):
    return ColumnMapping(columnName,ColumnMapping.ROBOT_TYPE_EQUIVALENCE)


class ColumnMapping:
    ROBOT_TYPE_ID = 1
    ROBOT_TYPE_LABEL = 2
    ROBOT_TYPE_PARENT = 3
    ROBOT_TYPE_RELATION = 4
    ROBOT_TYPE_ANNOTATION = 5
    ROBOT_TYPE_DISJOINT = 6
    ROBOT_TYPE_EQUIVALENCE = 7

    def __init__(self,excelColName,robotType,mappingId=None):
        self.excelColName = excelColName
        self.robotType = robotType
        self.mappingId = mappingId # relationship or annotation ID
        self.quoteNeeded = robotType in [
            ColumnMapping.ROBOT_TYPE_DISJOINT]

    def getRobotCodeString(self):
        if self.robotType == ColumnMapping.ROBOT_TYPE_ID:
            return "ID"
        elif self.robotType == ColumnMapping.ROBOT_TYPE_LABEL:
            return "LABEL"
        elif self.robotType == ColumnMapping.ROBOT_TYPE_PARENT:
            return "SC % SPLIT=;"
        elif self.robotType == ColumnMapping.ROBOT_TYPE_RELATION:
            return "SC "+self.mappingId+" some % SPLIT=;"
        elif self.robotType == ColumnMapping.ROBOT_TYPE_ANNOTATION:
            return "A "+self.mappingId+" SPLIT=;"
        elif self.robotType == ColumnMapping.ROBOT_TYPE_DISJOINT:
            return "DC % SPLIT=;"
        elif self.robotType == ColumnMapping.ROBOT_TYPE_EQUIVALENCE:
            return "EC %"

    def parseValue(self,value):
        if value is None:
            return ''
        else:
            value_a = value.encode("ascii","ignore").decode("utf-8")
            value_a = re.sub('\(.*?\)', '', value_a)
            value_a = re.sub('\[.*?\]', '', value_a)
            value_a = value_a.strip()
            if (self.quoteNeeded):
                values = value_a.split(';')
                values = [v.strip() for v in values]
                values = [quoteIfNeeded(v) for v in values]
                value_a = ";".join(values)
            return(value_a)

# Mappings need to exist for the name of each column in a template spreadsheet
HEADER_MAPPINGS = {"BCIO_ID": getIdMapping("BCIO_ID"),
    "ID": getIdMapping("ID"),
    "Name": getLabelMapping("Name"),
    "Label": getLabelMapping("Label"),
    "Label (synonym)": getLabelMapping("Label"),
    "Parent": getParentMapping("Parent"),
    "Parent class/ BFO class": getParentMapping("Parent"),
    "Logical definition": getEquivalenceMapping("Logical definition"),
    "Disjoint classes": getDisjointMapping("Disjoint classes"),
    "Definition":getAnnotationMapping("Definition","IAO:0000115"),
    "Definition_ID":getAnnotationMapping("Definition_ID","rdfs:isDefinedBy"),
    "Definition_Source":getAnnotationMapping("Definition_source","IAO:0000119"),
    "Definition source":getAnnotationMapping("Definition source","IAO:0000119"),
    "Examples":getAnnotationMapping("Examples","IAO:0000112"),
    "Examples of usage":getAnnotationMapping("Examples","IAO:0000112"),
    "Elaboration":getAnnotationMapping("Elaboration","IAO:0000112"),
    "Curator note":getAnnotationMapping("Curator note","IAO:0000232"),
    "Synonyms":getAnnotationMapping("Synonyms","IAO:0000118"),
    "Comment":getAnnotationMapping("Comment","rdfs:comment"),
    "Curation status":getAnnotationMapping("Curation status","IAO:0000078")
                }

# Unmapped headers that should not be part of the template
HEADERS_TO_IGNORE = ["Structure","BFO entity","Sub-ontology","Informal definition"]

# Provides a wrapper for easily executing common robot template functionality
# from within Python based on Excel spreadsheets. See https://github.com/ontodev/robot
class RobotWrapper:
    def __init__(self,robotcmd,cleanup=True):
        self.cleanup = cleanup
        self.robotcmd = robotcmd

    def __executeCommand__(self,command_str,shell_flag=True):
        Output = subprocess.Popen(command_str,
                shell=shell_flag,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT)
        stdout,stderr = Output.communicate()
        if stdout is not None and len(stdout) > 0:
            print(stdout)
        if stderr is not None:
            print(stderr)


class RobotImportsWrapper(RobotWrapper):

    # Handle externally imported content
    def processImportsFromExcel(self,importsFileName,
                                    importsOWLURI,importsOWLFileName,ontologyName):

        try:
            wb = openpyxl.load_workbook(importsFileName)
        except Exception as e:
            print(e)
            raise Exception("Not able to parse input file: "+importsFileName+":"+e)

        sheet = wb.active
        data = sheet.rows

        header = [i.value for i in next(data)[0:5]]

        formerge = []

        if not os.path.exists("temp"):
            os.mkdir("temp")

        for row in data:
            rowdata = [i.value for i in row[0:6]]
            onto_id = rowdata[0]
            purl = rowdata[1]
            root_id = rowdata[2]
            ids = rowdata[3]
            intermediates = rowdata[4]
            prefix = rowdata[5]

            if intermediates is None: intermediates = 'minimal'

            onto_shortname = purl[(purl.rindex('/')+1):]

            # Only download if we don't already have it.
            #Use cleanup=TRUE to clean up afterwards for fresh download next time.
            if not os.path.exists("temp/"+onto_shortname):
                get_ontology_cmd = 'curl -L '+purl+' > temp/' +onto_shortname
                self.__executeCommand__(get_ontology_cmd,shell_flag=True)

            root_id = root_id[root_id.find("[")+1:root_id.find("]")] # extract just the ID
            filename = 'temp/'+onto_id+'-slim.owl'
            slim_cmd = [self.robotcmd, 'merge',
                    '--input', 'temp/'+onto_shortname,
                    'extract', '--method', 'MIREOT',
                    '--annotate-with-source', 'true',
                    '--upper-term',root_id,
                    '--intermediates', intermediates,
                    '--output', filename ]
            if prefix:
                slim_cmd.append('--prefix')
                slim_cmd.append(prefix)
            terms = ids.split(";")
            for term_id in terms:
                term_id = term_id[term_id.find("[")+1:term_id.find("]")] # extract just the ID
                slim_cmd.append('--lower-term')
                slim_cmd.append(term_id)
            slim_cmd = " ".join(slim_cmd)
            print(slim_cmd)
            self.__executeCommand__(slim_cmd,shell_flag=True)
            formerge.append(filename)

        # Now merge all the imports into a single file
        merge_cmd = [self.robotcmd,'merge']
        for mergefile in formerge:
            merge_cmd.append('--input')
            merge_cmd.append(mergefile)
        merge_cmd.extend(['annotate', '--ontology-iri', importsOWLURI, '--version-iri', importsOWLURI, '--annotation rdfs:comment ','"This file contains externally imported content for the '+ontologyName+'. It was prepared using ROBOT and a custom script from a spreadsheet of imported terms."','--output', importsOWLFileName])
        merge_cmd = " ".join(merge_cmd)
        print(merge_cmd)
        self.__executeCommand__(merge_cmd,shell_flag=True)

        # Now delete the temp directory
        if (self.cleanup): shutil.rmtree('temp')

    def addAdditionalContent(self, extraContentTemplate,                                         importsOWLURI,importsOWLFileName):
        owlFileName = importsOWLURI[(importsOWLURI.rindex('/')+1):]
        owlTempFileName = owlFileName.replace(".owl","-temp.owl")

        robot_cmd = [self.robotcmd, 'template', '--template', extraContentTemplate,
                '--ontology-iri', importsOWLURI,
                '--output', owlTempFileName
                ]

        robot_cmd = " ".join(robot_cmd)

        print("About to execute Robot command: ",robot_cmd)

        self.__executeCommand__(command_str=robot_cmd)

        robot_cmd = [self.robotcmd, 'merge', '--input',owlFileName,'--input',owlTempFileName, '--output',owlFileName]

        robot_cmd = " ".join(robot_cmd)

        print("About to execute Robot command: ",robot_cmd)

        self.__executeCommand__(command_str=robot_cmd)


    # Remove metadata that causes a problem in Pronto
    # Overwrites original file so be careful
    def removeProblemMetadata(self, importsOWLURI, importsOWLFileName, metadataURIFile):

        robot_cmd = [self.robotcmd, 'remove', '--input', importsOWLFileName,
                '--term-file', metadataURIFile,
                '--axioms', 'annotation',
                '--output', importsOWLFileName
                ]

        robot_cmd = " ".join(robot_cmd)

        print("About to execute Robot command: ",robot_cmd)

        self.__executeCommand__(command_str=robot_cmd)



    # Also save an OBO (For Pronto)
    def createOBOFile(self, importsOWLURI, importsOWLFileName):
        owlFileName = importsOWLURI[(importsOWLURI.rindex('/')+1):]

        oboFileName = owlFileName.replace(".owl",".obo")

        robot_cmd = [self.robotcmd, 'merge', '--input', owlFileName, 'convert', '--output',oboFileName, '--check','false']

        robot_cmd = " ".join(robot_cmd)

        print("About to execute Robot command: ",robot_cmd)

        self.__executeCommand__(command_str=robot_cmd)





class RobotTemplateWrapper(RobotWrapper):

    def __init__(self,robotcmd):
        self.robotcmd = robotcmd
        self.cleanup = True
        self.all_entity_names = {}  # index of names to ontology entities
        self.all_entity_ids = {}    # index of ids to ontology entities
        self.headers_mapped = []   # Excel to CSV file headers mapping
        self.all_rel_names = {}  # index of names to relations
        self.all_rel_ids = {}    # index of ids to relations
        self.parents_to_children = {}


    def __dfs__(self,order, node):
        if node not in order.keys() and node.lower() in self.all_entity_names.keys():
            order[node] = ''
            if node in self.parents_to_children.keys():
                for child in self.parents_to_children[node]:
                    self.__dfs__(order, child)
        return order

    # Extract class information from an Excel spreadsheet
    def processClassInfoFromExcel(self, excelFileName):

        try:
            wb = openpyxl.load_workbook(excelFileName)
        except Exception as e:
            print(e)
            raise Exception("Error! Not able to parse file: "+excelFileName)

        sheet = wb.active
        data = sheet.rows

        header = [i.value for i in next(data)]
        print (header)

        # Check all header strings are in the header mapping or else fail with an error
        headers_notmapped = [h for h in header if h not in HEADER_MAPPINGS.keys()]
        print (headers_notmapped)

        for h in headers_notmapped:
            if h is not None and 'REL' in h:
                values = quoted.findall(h)
                if len(values) > 0:
                    HEADER_MAPPINGS[h]=getRelationshipMapping(h,relId=quoteIfNeeded(values[0]))

        headers_notmapped = [h for h in header if h not in HEADER_MAPPINGS.keys() and h is not None and h not in HEADERS_TO_IGNORE]

        if len(headers_notmapped) > 0:
            print("HEADERS NOT MAPPED: ",headers_notmapped, "IGNORING...")
            for h in headers_notmapped:
                HEADERS_TO_IGNORE.append(h)


        self.headers_mapped = [HEADER_MAPPINGS[h] for h in header if h is not None and h not in HEADERS_TO_IGNORE]
        header_indices = [i for i,h in zip(range(len(header)),header) if h is not None and h not in HEADERS_TO_IGNORE]

        # Process the rows, create a CSV template at the same time
        csvFileName = excelFileName.replace(".xlsx",".csv")

        csvfile = open(csvFileName, 'w', newline='')
        csv_writer = csv.writer(csvfile, delimiter=',',quotechar='\"', quoting=csv.QUOTE_MINIMAL)

        csv_writer.writerow([header[i] for i in header_indices])
        csv_writer.writerow([c.getRobotCodeString() for c in self.headers_mapped])

        for row in data:
            row = [row[i] for i in header_indices] # just those headers that are mapped
            row = [r.value for r in row]
            l = [mapping.parseValue(i) for (i,mapping) in zip(row,self.headers_mapped)]

            entity = OntologyEntity()
            # Now also process and store the values for merging if needed
            for col_val,header_val in zip(row,self.headers_mapped):
                if header_val.robotType == ColumnMapping.ROBOT_TYPE_ID:
                    id = col_val
                    entity.id = id
                    self.all_entity_ids[id] = entity
                if header_val.robotType == ColumnMapping.ROBOT_TYPE_LABEL:
                    name = col_val
                    if name is None: continue
                    entity.name = name
                    entity.synonyms = []
                    self.all_entity_names[name.lower()] = entity
                    synonym = ''
                    if '(' in name and ')' in name:
                        synonym = re.search(r'\((.*?)\)',name).group(1)
                        if len(synonym) > 0:
                            name = name.split("(",1)[0].strip()
                            entity.name = name
                            entity.synonyms = [synonym]
                            self.all_entity_names[name.lower()] = entity
                            self.all_entity_names[synonym.lower()] = entity
                if header_val.excelColName == "Synonyms" and col_val is not None:
                    more_synonyms = col_val.split(";")
                    entity.synonyms.extend(more_synonyms)
                    for synonym in more_synonyms:
                        self.all_entity_names[synonym.lower().strip()] = entity
                if header_val.excelColName == "Definition":
                    entity.definition = col_val
                if header_val.excelColName == "Parent":
                    parents = col_val
                    if parents is not None:
                        parent = parents.split("/")[0]
                        if '(' in parent:
                            parent = parent.split("(")[0].strip()
                        if '[' in parent:
                            parent = parent.split("[")[0].strip()
                        entity.parent = parent
                if header_val.excelColName == "Examples":
                    examples = col_val
                    if examples is not None and len(examples) > 0:
                        entity.examples = examples
                if header_val.excelColName == "Comment":
                    comment = col_val
                    if comment is not None and len(comment) > 0:
                        entity.comment = comment
                if header_val.excelColName == "Curation status":
                    entity.curationStatus = col_val

            if entity.curationStatus not in ['Obsolete']:
                csv_writer.writerow(l)
            else:
                print("Not writing row to template due to obsolete status",entity.name)

        # close the csv file
        csvfile.close()
        print ('FINISHED PARSING ALL ROWS IN SPREADSHEET')
        wb.close()
        return csvFileName

    def processRelInfoFromExcel(self,excelFileName):

        try:
            wb = openpyxl.load_workbook(excelFileName)
        except Exception as e:
            print(e)
            raise Exception("Not able to parse input file: "+excelFileName)

        sheet = wb.active
        data = sheet.rows

        header = [i.value for i in next(data)[0:7]]
        print(header)

        for row in data:
            rowdata = [i.value for i in row[0:7]]
            id = rowdata[0]
            name = rowdata[1]
            if name is None: continue
            entity = OntologyRelation(id,name)
            entity.equivalent = rowdata[2]
            entity.parent = rowdata[3]
            entity.definition = rowdata[4]
            domain = rowdata[5]
            range = rowdata[6]
            self.all_rel_names[name.lower()] = entity
            self.all_rel_ids[id] = entity

        wb.close()


    def createCsvRelationTemplateFile(self,csvFileName):
        # Create ROBOT template for NEW properties (parent is not None)

        with open(csvFileName,'w', newline='') as rel_create_csv:
            rel_writer = csv.writer(rel_create_csv,delimiter=',',quotechar='\"', quoting=csv.QUOTE_MINIMAL)
            rel_writer.writerow(["Id","Name","Type","Parent","Def","Domain","Range"])
            rel_writer.writerow(["ID","LABEL","TYPE","SP %","A IAO:0000115","DOMAIN","RANGE"])

            for rel in self.all_rel_names.values():
                if rel.parent is not None:
                    parent = rel.parent[rel.parent.find("[")+1:rel.parent.find("]")] # extract just the ID
                    rel_writer.writerow([rel.id,rel.name,"object property",parent,rel.domain,rel.range])
                else:
                    rel_writer.writerow([rel.id,rel.name,"object property",None,rel.domain,rel.range])
        print("Finished writing relation creation template")


    def mergeRelInfoFromLucidChart(self,entities,relations):
        # Merge lucidchart information with definitions information to populate relations
        #for entity in entities.values():
        #    entity_name = entity.name.replace('\n',' ') # just the name matters
        #    # do we have this entity name in our all_entities?
        #    if entity_name.lower().strip() not in self.all_entity_names.keys():
        #        print("ERROR: Name ", entity_name, "NOT FOUND.")
        #        print(self.all_entity_names.keys())

        for rel in relations:
            rel_name = rel.relType
            if '(' in rel_name:
                rel_name = rel_name[0:rel_name.rindex('(')].strip()
            if rel_name.lower() not in self.all_rel_names.keys():
                print("ERROR: Rel ", rel_name, "NOT FOUND.")
                print(self.all_rel_names.keys())
                continue
            onto_rel = self.all_rel_names[rel_name.lower()]
            onto_entity1 = self.all_entity_names[rel.entity1.name.replace('\n',' ').lower().strip()]
            onto_entity2 = self.all_entity_names[rel.entity2.name.replace('\n',' ').lower().strip()]
            if onto_entity1.relations is None:
                onto_entity1.relations = {}
            if onto_rel.name not in onto_entity1.relations.keys():
                onto_entity1.relations[onto_rel.name] = []
            onto_entity1.relations[onto_rel.name].append(onto_entity2)


    def createMergedSpreadsheet(self,excelFileName):

        book = Workbook()
        sheet = book.active

        rel_header_ids = ["REL '"+s.name+"' ["+s.id+"]" for s in self.all_rel_names.values()]
        header = ('BCIO_ID','Name','Parent','Definition','Synonyms','Examples',*rel_header_ids)

        sheet.append(header)

        # PARENT classes AND TARGETS OF RELATIONS -- prepare list of required imports for information and cross-checking
        import_classes = []
        top_level = []
        for entity in self.all_entity_names.values():
            parent = entity.parent
            if parent.lower() not in self.all_entity_names.keys() and parent.lower() not in import_classes:
                import_classes.append(parent.lower())
                top_level.append(entity.name)
            if parent not in self.parents_to_children:
                self.parents_to_children[parent] = []
            self.parents_to_children[parent].append(entity.name)

        print(import_classes)

        order = {}
        for t in top_level:
            order1 = self.__dfs__({},t)
            order.update(order1)
        for t in self.parents_to_children.keys():
            order1 = self.__dfs__({},t)
            order.update(order1)
        order.update({x.name:'' for x in self.all_entity_ids.values()})
        order = order.keys()

        for entity_name in order:
            if entity_name is None:
                continue
            entity = self.all_entity_names[entity_name.lower()]
            parent_name = entity.parent.lower() if entity.parent.lower() in import_classes else self.all_entity_names[entity.parent.lower()].name
            rel_vals = [";".join([z.name for z in x]) if len(x)>0 else '' for x in [entity.relations[y.name] if entity.relations is not None and y.name in entity.relations else [] for y in self.all_rel_names.values()]]
            #print([y.name for y in self.all_rel_names.values()])
            #print(entity_name,rel_vals)
            row = (entity.id,
                entity.name,
                parent_name,
                entity.definition,
                ";".join(entity.synonyms),
                entity.examples,
                *rel_vals
            )
            sheet.append(row)

        book.save(excelFileName)


    # Executes ROBOT from a template file as created
    def createOntologyFromTemplateFile(self, csvFileName, dependency, iri_prefix, id_prefixes, ontology_iri,owlFileName):
        robot_cmd = [self.robotcmd, 'template', '--template', csvFileName]
        for p in id_prefixes:
            robot_cmd.append('--prefix')
            robot_cmd.append(p)
        robot_cmd.extend(   ['--ontology-iri', ontology_iri,
                '--output', owlFileName
                ] )

        # A bit of hacking to deal appropriately with external dependency files:
        if dependency is not None:
            # Allow multiple dependencies. These will become OWL imports.
            dependencyFileNames = dependency.split(',')
            print(dependencyFileNames)
            dependencyFileName = "imports.owl"
            with open(dependencyFileName,'w') as outFile:
                outFile.write("""<?xml version=\"1.0\"?>
    <rdf:RDF xmlns="http://www.semanticweb.org/ontologies/temporary#"
        xml:base="http://www.semanticweb.org/ontologies/temporary"
        xmlns:dc="http://purl.org/dc/elements/1.1/"
        xmlns:obo="http://purl.obolibrary.org/obo/"
        xmlns:owl="http://www.w3.org/2002/07/owl#"
        xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#"
        xmlns:xml="http://www.w3.org/XML/1998/namespace"
        xmlns:xsd="http://www.w3.org/2001/XMLSchema#"
        xmlns:foaf="http://xmlns.com/foaf/0.1/"
        xmlns:rdfs="http://www.w3.org/2000/01/rdf-schema#">
        <owl:Ontology rdf:about=\"""" + ontology_iri + "\">\n")
                for d in dependencyFileNames:
                    outFile.write("<owl:imports rdf:resource=\""+iri_prefix+ d +"\"/> \n")
                outFile.write(" </owl:Ontology> \n</rdf:RDF> ")

        robot_cmd.extend(['--input',dependencyFileName,"--merge-before","--collapse-import-closure","false"])

        robot_cmd = " ".join(robot_cmd)

        print("About to execute Robot command: ",robot_cmd)

        self.__executeCommand__(command_str=robot_cmd)




class RobotSubsetWrapper(RobotWrapper):

    def __init__(self,robotcmd):
        self.robotcmd = robotcmd

    def createSubsetFrom(self, inputOntologyFileName, outputFileName, rootId, idPrefix, exportCsvHeaders=None,exportSort=None):
        robot_cmd = [self.robotcmd,'merge',
        '--input', inputOntologyFileName,
         'extract',
        '--method', 'MIREOT',
        '--prefix', idPrefix,
        '--annotate-with-source', 'true',
        '--branch-from-term',rootId,
        '--intermediates', 'all']

        if exportCsvHeaders:
            robot_cmd.extend(['export','--header','"'+exportCsvHeaders+'"',
                            '--prefix', idPrefix,
                            '--split', '"; "',
                            '--export',outputFileName])
            if exportSort:
                robot_cmd.extend(['--sort','"'+exportSort+'"'])
        else:
            robot_cmd.extend( ['--output', outputFileName ] )

        robot_cmd = " ".join(robot_cmd)

        print("About to execute Robot command: ",robot_cmd)

        self.__executeCommand__(command_str=robot_cmd)
